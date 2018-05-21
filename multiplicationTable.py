#! python3
# multiplicatoinTable.py    - Takes a number for argument line and generates
# NxN multiplication table for it

import openpyxl, os
from openpyxl.utils import get_column_letter

#os.chdir('C:\\chapter12')

print('Please give a number to generate a multiplication table for: ')

number = input()

print('Generating a muliplication table for %i' %(int(number)))

wb = openpyxl.Workbook()
sheet = wb.active

# Generate rows and columns




for i in range(1, int(number) + 2):
    sheet['A1'] = None
    sheet['A' + str(i)] = i - 1
    sheet[get_column_letter(i) + str(1)] = i - 1

for column in range(2, int(number) + 2):
    letter = get_column_letter(column)

    for rows in range(2, int(number) + 2):
        sheet[letter + str(rows)] = sheet['A' + str(rows)].value * sheet[letter + '1'].value


wb.save('MultiplicatoinTable.xlsx')
print('Dana')
