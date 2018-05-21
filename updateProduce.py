#! python3
# updateProduce.py  - looks thorugh excel sheet and fides a specific produce and
# updates its price

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

priceUpdates = {'Garlic': 3.07,
                'Celery': 1.19,
                'Lemon': 1.27}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in priceUpdates:
        sheet.cell(row=rowNum, column=2).value = priceUpdates[produceName]

        
wb.save('updatedProduceSales.xlsx')

